import os
import glob
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import math

# Define file paths
output_folder = "C:/Users/Wanderer/Documents/OSU-GT-STANFORD/COBRA.UNIT/!README/pqr-backend-output-files"
final_output_file = os.path.join(output_folder, "pqrFinalOutput.xlsx")

# List of pass/fail headers in the desired order
passFailHeaders = [
    "grossMarginCheck1",
    "grossMarginCheck2",
    "netProfitMarginCheck1",
    "netProfitMarginCheck2",
    "peTimesPriceToBookRatioCheck",
    "marketCapCheck",
    "currentRatioCheck",
    "ROACheck",
    "returnOnEquityCheck",
    "ROICCheck",
    "researchAndDevelopmentCheck",
    "incomeTaxExpensesCheck",
    "betaCheck",
    "priceVolumeCheck",
    "earningsYieldCheck"
]

# Function to get all pqrOutput files (sorted)
def get_output_files():
    return sorted(glob.glob(os.path.join(output_folder, "pqrOutput*.csv")))

# Step 1: Collect all unique headers and move debtToEquity before rewriting for passFailHeaders
def get_master_header(output_files):
    master_header = set()

    for file in output_files:
        with open(file, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            header = next(reader, None)
            if header:
                master_header.update(header)

    master_header = list(sorted(master_header))  # Convert back to a sorted list

    # Ensure "Ticker" is always first
    if "Ticker" in master_header:
        master_header.remove("Ticker")
    ordered_header = ["Ticker"]

    # Ensure "totalPasses" is last
    if "totalPasses" in master_header:
        master_header.remove("totalPasses")

    # Ensure "debtToEquity" is at the end of data columns before passFailHeaders
    if "debtToEquity" in master_header:
        master_header.remove("debtToEquity")

    # Separate headers into groups
    data_headers = [col for col in master_header if col not in passFailHeaders and col != "totalPasses"]
    pass_fail_headers_in_master = [col for col in passFailHeaders if col in master_header]

    # Reconstruct header order: Ticker -> Data -> debtToEquity -> Pass/Fail -> totalPasses -> setSellPrice120% -> newPeTimesPriceToBookRatio
    ordered_header.extend(data_headers)  # Add all data columns first
    ordered_header.append("debtToEquity")  # Append debtToEquity at the end of data columns
    ordered_header.extend(pass_fail_headers_in_master)  # Then add passFailHeaders
    ordered_header.append("totalPasses")  # Ensure totalPasses is last
    ordered_header.append("setSellPrice120%")  # Add new column at the end
    ordered_header.append("newPeTimesPriceToBookRatio")  # Add new column after setSellPrice120%
    ordered_header.append("maxSellPrice")  # New calculated column
    ordered_header.append("totalPercentGainMaxSellPrice")  # New calculated column

    return ordered_header

# Helper function to safely convert numbers
def parse_number(value):
    try:
        return float(value)
    except (ValueError, TypeError):
        return float('-inf')  

# Step 2: Align, merge, filter, and sort data by "totalPasses" in descending order
def align_and_write_data(output_files, master_header):
    all_rows = []  # Store valid rows

    for file in output_files:
        print(f"Processing file: {file}")

        with open(file, 'r', newline='', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)

            if not rows:
                continue

            file_header = rows[0]  
            data_rows = rows[1:]  

            # Create a mapping of column indices for current file
            header_index_map = {col: idx for idx, col in enumerate(file_header)}

            # Ensure necessary indices exist
            required_indices = ["grossMarginCheck1", "netProfitMarginCheck1", "peTimesPriceToBookRatioCheck", 
                                "returnOnEquityCheck", "ROICCheck", "betaCheck", "debtToEquity", "stockPrice", "peRatio", "priceToBookRatio"]
            column_indices = {col: master_header.index(col) for col in required_indices if col in master_header}

            # Explicitly get indices for new columns
            set_sell_price_index = master_header.index("setSellPrice120%")
            new_pe_pb_ratio_index = master_header.index("newPeTimesPriceToBookRatio")
            max_sell_price_index = master_header.index("maxSellPrice")
            total_percent_gain_max_sell_price_index = master_header.index("totalPercentGainMaxSellPrice")

            # Process rows, filtering invalid ones
            for row in data_rows:
                aligned_row = [''] * len(master_header)  # Initialize full row

                # Populate row with aligned data
                for col_name, idx in header_index_map.items():
                    if col_name in master_header and idx < len(row):
                        aligned_row[master_header.index(col_name)] = row[idx]

                # Convert values for filtering
                pe_value = parse_number(aligned_row[column_indices.get("peTimesPriceToBookRatioCheck", -1)])
                gm_value = parse_number(aligned_row[column_indices.get("grossMarginCheck1", -1)])
                np_value = parse_number(aligned_row[column_indices.get("netProfitMarginCheck1", -1)])
                roe_value = parse_number(aligned_row[column_indices.get("returnOnEquityCheck", -1)])
                roic_value = parse_number(aligned_row[column_indices.get("ROICCheck", -1)])
                beta_value = parse_number(aligned_row[column_indices.get("betaCheck", -1)])

                # DELETE ROWS BASED ON CONDITIONS
                if (
                    pe_value == 0 or 
                    (gm_value == 0 and np_value == 0) or 
                    (roe_value == 0 and beta_value == 0) or 
                    np_value == 0 or  # Remove if netProfitMarginCheck1 is 0
                    roe_value == 0 or  # Remove if returnOnEquityCheck is 0
                    roic_value == 0  # Remove if ROICCheck is 0
                ):
                    continue  # Skip row if it meets any filtering conditions

                # Calculate new values
                stock_price = parse_number(aligned_row[column_indices.get("stockPrice", -1)])
                pe_ratio = parse_number(aligned_row[column_indices.get("peRatio", -1)])
                pb_ratio = parse_number(aligned_row[column_indices.get("priceToBookRatio", -1)])

                set_sell_price_120 = stock_price * 1.2 if stock_price > 0 else ""
                temp_pe_ratio = pe_ratio * 1.2 if pe_ratio > 0 else ""
                temp_pb_ratio = pb_ratio * 1.2 if pb_ratio > 0 else ""
                new_pe_pb_ratio = temp_pe_ratio * temp_pb_ratio if temp_pe_ratio and temp_pb_ratio else ""

                # Skip rows where newPeTimesPriceToBookRatio > 22.5
                if new_pe_pb_ratio and new_pe_pb_ratio > 22.5:
                    continue


# Calculate tempPeTimesPbRatio
                temp_pe_pb_ratio = pe_ratio * pb_ratio if pe_ratio > 0 and pb_ratio > 0 else None

                # Calculate maxMultiplier (formerly whatNumber)
                max_multiplier = math.sqrt(22.5 / temp_pe_pb_ratio) if temp_pe_pb_ratio and temp_pe_pb_ratio > 0 else None

                # Calculate maxPercentGain
                max_percent_gain = max_multiplier if max_multiplier else None

                # Calculate maxSellPrice
                max_sell_price = stock_price * max_percent_gain if max_percent_gain and stock_price > 0 else None

                # Calculate totalPercentGainMaxSellPrice
                total_percent_gain_max_sell_price = ((max_sell_price - stock_price) / stock_price) * 100 if max_sell_price and stock_price > 0 else None

                # Correctly assign new values in their respective positions in aligned_row
                aligned_row[set_sell_price_index] = set_sell_price_120
                aligned_row[new_pe_pb_ratio_index] = new_pe_pb_ratio
                aligned_row[max_sell_price_index] = max_sell_price
                aligned_row[total_percent_gain_max_sell_price_index] = total_percent_gain_max_sell_price

                all_rows.append(aligned_row)  

    # Sort rows by totalPasses (handle negatives and non-numeric values correctly)
    total_passes_index = master_header.index("totalPasses")
    all_rows.sort(key=lambda x: parse_number(x[total_passes_index]), reverse=True)

    # Write final output as an Excel file with red highlighting
    wb = Workbook()
    ws = wb.active
    ws.append(master_header)  # Write headers

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for row in all_rows:
        ws.append(row)
        gm_col_index = master_header.index("grossMarginCheck1") + 1  # Excel is 1-based index
        beta_col_index = master_header.index("betaCheck") + 1  # Excel is 1-based index

        # Highlight grossMarginCheck1 cell in red if it equals 0
        if parse_number(row[gm_col_index - 1]) == 0:
            ws.cell(row=ws.max_row, column=gm_col_index).fill = red_fill

        # Highlight betaCheck cell in green if it equals 1
        if parse_number(row[beta_col_index - 1]) == 1:
            ws.cell(row=ws.max_row, column=beta_col_index).fill = green_fill


    wb.save(final_output_file)
    print(f"Data successfully written to {final_output_file}")

if __name__ == "__main__":
    output_files = get_output_files()

    if not output_files:
        print("No files found.")
    else:
        master_header = get_master_header(output_files)
        align_and_write_data(output_files, master_header)
